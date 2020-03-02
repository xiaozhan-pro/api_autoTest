"""
excel类封装需要提供以下功能：

1、打开表单

2、读取标题

3、读取所有的数据

4、指定单元格写入数据（使用静态方法，不要使用实例方法）
"""
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

class ExcelHandler():
    def __init__(self,path):
        self.path=path

    def open_sheet(self,name) ->Worksheet:
        """
        打开表单
        在函数或方法后面加 -> 表示返回值是这样一个类型
        这里加的好处在于解决 "通过sheet名字，wb["sheet_name"]，加了.无提示" 这个问题
        """
        wb = openpyxl.load_workbook(self.path)
        sheet =wb[name]
        return sheet

    def sheet_header(self,sheet_name) :
        """读取标题"""
        sheet=self.open_sheet(sheet_name)
        header=[]
        for data in sheet[1]:
            header.append(data.value)
        return header

    def sheet_readAll(self,sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        header=[]
        for data in sheet[1]:
            header.append(data.value)  #读取标题

        total_rows=list(sheet.rows)[1:]
        data_all=[]
        for row_data in total_rows:
            data=[]
            for cell in row_data:
                data.append(cell.value)
                data_dict=dict(zip(header,data))
            data_all.append(data_dict)

        return data_all

    @staticmethod
    def sheet_writeCell(path,sheet_name,row,column,value):
        """
        指定单元格写入数据
        1、指定行和列写入值
        2、保存表格
        3、关闭表格
        """
        wb = openpyxl.load_workbook(path)
        sheet =wb[sheet_name]
        sheet.cell(row, column) .value = value
        wb.save(path)
        wb.close()




